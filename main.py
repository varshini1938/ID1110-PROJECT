import openpyxl
#openpyxl is a library which contains functions to perform various function to an excel sheet such as opening,editing, etc.
import smtplib
#smtplib is used to send mails 
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
#these are used to write the message in the mail

#opening the excel sheet
book = openpyxl.load_workbook('C:\\Users\\smtca\\Pictures\\Sri\\ACADEMICS\\Python\\project\\attendance.xlsx')
#load_workbook function opens and reads files of the format .xlsx

# opening the required sheet
sheet = book['attendance']

# number of students are the number of rows-1
rows = sheet.max_row - 1

# number of subjects are number of number columns-2
sub = sheet.max_column - 2

# list of students who have 1 leave left(they have taken k-1 leaves if k is the max number of leaves that can be taken)
l1 = []

#l2 concatenates the roll no.'s of the students that have taken the maximum no.of leaves
l2 = ""

#list of students who have taken maximum no.of leaves
l3 = []

#staff mail ids
staff_mails = ['srivarshini4578@gmail.com', '132301012@smail.iitpkd.ac.in', '132301012@smail.iitpkd.ac.in']

# Warning messages
message1 = "Warning!!! you can take only one more day leave for Physics class"
message2 = "Warning!!! you can take only one more day leave for Math class"
message3 = "Warning!!! you can take only one more day leave for Mechanics class"
message4 = "Warning!!! you can take only one more day leave for Python class"

#creating fuction to update the file
def savefile():
    book.save(r'C:\\Users\\smtca\\Pictures\\Sri\\ACADEMICS\\Python\\project\\attendance.xlsx')
    print("saved!")

#taking input of the maximum number of leaves that can be taken by a student
k = int(input('enter the maximum number of leaves that can be taken: '))

#creating a fuction to check the number of leaves taken
def check(leaves, row_num, b):
    # to globally use the lists and strings
    global staff_mails
    global l2
    global l3

    for n in range(0, len(row_num)):
        # if the student has taken k-1 leaves
        if leaves[n] == k-1:
            if b == 1:
                # b is the subject code
                #appending the mail id to list1
                l1.append(sheet.cell(row=row_num[n], column=2).value)
                mailstu(l1, message1)
            elif b == 2:
                l1.append(sheet.cell(row=row_num[n], column=2).value)
                mailstu(l1, message2)
            elif b == 3:
                l1.append(sheet.cell(row=row_num[n], column=2).value)
                mailstu(l1, message3)
            else:
                l1.append(sheet.cell(row=row_num[n], column=2).value)
                mailstu(l1, message4)
        
        # if the student has taken thr maximum number of leaves or more
        elif leaves[n] > k-1:
            if b == 1:
                #adding the roll no. to l2
                l2 = l2 + str(sheet.cell(row=row_num[n], column=1).value)
                #adding the mail id to list3
                l3.append(sheet.cell(row=row_num[n], column=2).value)
                subject = "Physics"
            elif b == 2:
                l2 = l2 + str(sheet.cell(row=row_num[n], column=1).value)
                l3.append(sheet.cell(row=row_num[n], column=2).value)
                subject = "Math"
            elif b == 3:
                l2 = l2 + str(sheet.cell(row=row_num[n], column=1).value)
                l3.append(sheet.cell(row=row_num[n], column=2).value)
                subject = "Mechanics"
            else:
                l2 = l2 + str(sheet.cell(row=row_num[n], column=1).value)
                l3.append(sheet.cell(row=row_num[n], column=2).value)
                subject = "Python"

        # this message is sent to the students who have taken maximum no.of leaves
        if l2 != "" and len(l3) != 0:
            # message for student
            msg1 = "You have lack of attendance in " + subject + " !!!"
            # message for staff
            msg2 = "The following students have lack of attendance in your subject : " + l2
            mailstu(l3, msg1) # function to send mail to students
            staff_id = staff_mails[b-1]# choosing the respective staff mail id
            mailstaff(staff_id, msg2)# function to send a mail to the staff

# for students
def mailstu(li, msg):
    from_id = '132301012@smail.iitpkd.ac.in'
    pwd = 'Ts07fm4578!'
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)
   
    # for each student to warn send mail
    for i in range(0, len(li)):
        to_id = li[i]
        message = MIMEMultipart()
        message['Subject'] = 'Attendance report'
        message.attach(MIMEText(msg, 'plain'))
        content = message.as_string()
        s.sendmail(from_id, to_id, content)
    s.quit()
    print("mail sent to students")

# for staff
def mailstaff(mail_id, msg):
    from_id = '132301012@smail.iitpkd.ac.in'
    pwd = 'Ts07fm4578!'
    to_id = mail_id
    message = MIMEMultipart()
    message['Subject'] = 'Lack of attendance report'
    message.attach(MIMEText(msg, 'plain'))
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)
    content = message.as_string()
    s.sendmail(from_id, to_id, content)
    s.quit()
    print('Mail Sent to staff')

# variable for looping for input
resp = 1

#updating the attendance to the excel sheet
while resp == 1:
    #subject numbers
    print("1--->phy\n2--->math\n3--->mech\n4--->python")

    # enter the corresponding number
    t = int(input("enter subject :"))

    # taking input of the number of people absent for that course
    no_of_absentees = int(input('no.of.absentees :'))

    #taking the roll no.'s of the students on leave 
    if no_of_absentees > 1:
        x = list(map(int, (input('roll nos :').split(','))))
    else:
        x = [int(input('roll no :'))]

    # list to hold row of the student in the excel sheet
    row_num = []

    # list to hold the total number of leaves taken by a particular student
    leaves = []

    #updating the excel sheet
    for n in x:
        #students
        for i in range(2, rows+2):
            if t == 1:
                if sheet.cell(row=i, column=1).value == n:
                    #updating the number of leaves
                    s = sheet.cell(row=i, column=3).value
                    s = s + 1
                    sheet.cell(row=i, column=3).value = s
                    #saving the data
                    savefile()
                    leaves.append(s)
                    row_num.append(i)
            elif t == 2:
                if sheet.cell(row=i, column=1).value == n:
                    s = sheet.cell(row=i, column=4).value
                    s = s + 1
                    sheet.cell(row=i, column=4).value = s
                    savefile()
                    leaves.append(s)
                    row_num.append(i)
            elif t == 3:
                if sheet.cell(row=i, column=1).value == n:
                    s = sheet.cell(row=i, column=5).value
                    s = s + 1
                    sheet.cell(row=i, column=5).value = s
                    savefile()
                    leaves.append(s)
                    row_num.append(i)
            elif t == 4:
                if sheet.cell(row=i, column=1).value == n:
                    s = sheet.cell(row=i, column=6).value
                    s = s + 1
                    sheet.cell(row=i, column=6).value = s
                    savefile()
                    leaves.append(s)
                    row_num.append(i)

    check(leaves, row_num, t)
    #taking the input if the user wants to check for another subject
    resp = int(input('another subject ? 1---->yes 0--->no'))

